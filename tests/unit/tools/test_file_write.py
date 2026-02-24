"""Tests for file_write tool."""

from __future__ import annotations

import csv
import io
import json
from pathlib import Path

from zhi.tools.file_write import FileWriteTool


class TestFileWriteText:
    def test_write_text_file(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="hello.txt", content="Hello!")
        assert "File written" in result
        assert (out / "hello.txt").read_text(encoding="utf-8") == "Hello!"

    def test_write_creates_output_dir(self, tmp_path: Path) -> None:
        out = tmp_path / "new_output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="test.txt", content="data")
        assert "File written" in result
        assert out.exists()
        assert (out / "test.txt").exists()


class TestFileWriteMarkdown:
    def test_write_markdown(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="readme.md", content="# Title\n\nContent here.")
        assert "File written" in result
        assert "# Title" in (out / "readme.md").read_text(encoding="utf-8")


class TestFileWriteJson:
    def test_write_json_dict(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        data = {"key": "value", "number": 42}
        result = tool.execute(path="data.json", content=data)
        assert "File written" in result
        loaded = json.loads((out / "data.json").read_text(encoding="utf-8"))
        assert loaded["key"] == "value"
        assert loaded["number"] == 42

    def test_write_json_string(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="data.json", content='{"a": 1}')
        assert "File written" in result
        loaded = json.loads((out / "data.json").read_text(encoding="utf-8"))
        assert loaded["a"] == 1


class TestFileWriteCsv:
    def test_write_csv(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        data = {"headers": ["Name", "Age"], "rows": [["Alice", 30], ["Bob", 25]]}
        result = tool.execute(path="people.csv", content=data)
        assert "File written" in result

        content = (out / "people.csv").read_text(encoding="utf-8")
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
        assert rows[0] == ["Name", "Age"]
        assert rows[1] == ["Alice", "30"]


class TestFileWriteXlsx:
    def test_write_xlsx(self, tmp_path: Path) -> None:
        import openpyxl

        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        data = {
            "sheets": [
                {
                    "name": "People",
                    "headers": ["Name", "Age"],
                    "rows": [["Alice", 30], ["Bob", 25]],
                }
            ]
        }
        result = tool.execute(path="people.xlsx", content=data)
        assert "File written" in result

        wb = openpyxl.load_workbook(out / "people.xlsx")
        ws = wb.active
        assert ws.title == "People"
        assert ws.cell(1, 1).value == "Name"
        assert ws.cell(2, 1).value == "Alice"

    def test_write_xlsx_multiple_sheets(self, tmp_path: Path) -> None:
        import openpyxl

        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        data = {
            "sheets": [
                {"name": "Sheet1", "headers": ["A"], "rows": [["val1"]]},
                {"name": "Sheet2", "headers": ["B"], "rows": [["val2"]]},
            ]
        }
        result = tool.execute(path="multi.xlsx", content=data)
        assert "File written" in result

        wb = openpyxl.load_workbook(out / "multi.xlsx")
        assert len(wb.sheetnames) == 2
        assert wb.sheetnames[0] == "Sheet1"
        assert wb.sheetnames[1] == "Sheet2"


class TestFileWriteDocx:
    def test_write_docx(self, tmp_path: Path) -> None:
        import docx

        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        data = {"content": "# Title\n\nA paragraph."}
        result = tool.execute(path="doc.docx", content=data)
        assert "File written" in result

        doc = docx.Document(str(out / "doc.docx"))
        texts = [p.text for p in doc.paragraphs]
        assert "Title" in texts[0]
        assert "A paragraph." in texts[1]


class TestFileWriteOutputDir:
    def test_writes_to_output_dir(self, tmp_path: Path) -> None:
        out = tmp_path / "zhi-output"
        tool = FileWriteTool(output_dir=out)
        tool.execute(path="test.txt", content="data")
        assert (out / "test.txt").exists()


class TestFileWriteCreatesDir:
    def test_creates_subdirectory(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="sub/deep/file.txt", content="nested")
        assert "File written" in result
        assert (out / "sub" / "deep" / "file.txt").exists()


class TestFileWriteNoOverwrite:
    def test_refuses_overwrite(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        out.mkdir()
        (out / "existing.txt").write_text("old", encoding="utf-8")
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="existing.txt", content="new")
        assert "Error" in result
        assert "already exists" in result
        # Original content unchanged
        assert (out / "existing.txt").read_text(encoding="utf-8") == "old"


class TestFileWritePathTraversal:
    def test_rejects_dotdot(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="../escape.txt", content="evil")
        assert "Error" in result
        assert "traversal" in result.lower()

    def test_rejects_absolute_path(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="/tmp/evil.txt", content="evil")
        assert "Error" in result
        assert "Absolute" in result


class TestFileWriteSymlinkAttackBlocked:
    def test_symlink_outside_output_blocked(self, tmp_path: Path) -> None:
        out = tmp_path / "output"
        out.mkdir()
        external = tmp_path / "external"
        external.mkdir()
        # Create a symlink inside output that points outside
        link = out / "escape"
        link.symlink_to(external)

        tool = FileWriteTool(output_dir=out)
        result = tool.execute(path="escape/evil.txt", content="pwned")
        assert "Error" in result
        assert "outside" in result.lower() or "symlink" in result.lower()


class TestFileWriteRiskyFlag:
    def test_file_write_is_risky(self) -> None:
        tool = FileWriteTool()
        assert tool.risky is True


class TestFileWriteErrorCases:
    def test_missing_path(self, tmp_path: Path) -> None:
        tool = FileWriteTool(output_dir=tmp_path)
        result = tool.execute(content="data")
        assert "Error" in result

    def test_missing_content(self, tmp_path: Path) -> None:
        tool = FileWriteTool(output_dir=tmp_path)
        result = tool.execute(path="test.txt")
        assert "Error" in result


class TestFileWriteCrossPlatform:
    def test_path_validation_cross_platform(self, tmp_path: Path) -> None:
        """Ensure path check works on all platforms (no hard-coded '/')."""
        tool = FileWriteTool(output_dir=tmp_path)
        result = tool.execute(path="test.txt", content="hello")
        assert "File written" in result
        assert "Error" not in result
